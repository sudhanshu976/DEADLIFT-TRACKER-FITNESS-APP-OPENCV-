# Importing libraries
import cv2
import mediapipe as mp
import numpy as np
import time
import tkinter as tk
from tkinter import simpledialog
from datetime import datetime
import os.path
from openpyxl import Workbook, load_workbook

# Get the display resolution
screen_width = 1080
screen_height = 720

# Storing main utilities
mp_drawing = mp.solutions.drawing_utils
mp_pose = mp.solutions.pose

# Function for calculating angles between 3 points
def calculateAngle(a, b, c):
    a = np.array(a)  # First Point
    b = np.array(b)  # Second Point
    c = np.array(c)  # Last Point
    radians = np.arctan2(c[1] - b[1], c[0] - b[0]) - np.arctan2(a[1] - b[1], a[0] - b[0])
    angle = np.abs(radians * 180.0 / np.pi)

    if angle > 180.0:
        angle = 360 - angle
    return angle

#Initializing variable as it was causing error
shoulder=None

# Function to calculate and display angle between two points
def calculateAndDisplayAngle(frame, shoulder, hip):
    angle = calculateAngle(shoulder, hip, [shoulder[0], hip[1]])
    if angle < 70:
        cv2.putText(frame, f"FORM: BAD ", (10, 190), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2, cv2.LINE_AA)
    else:
        cv2.putText(frame, f"FORM: GOOD ", (10, 190), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2, cv2.LINE_AA)

# Function to save report data to Excel
def save_report_to_excel(counter, total_time, rep_times):
    # Open the existing Excel file or create a new one
    excel_file = "reports_summary.xlsx"
    wb = Workbook() if not os.path.isfile(excel_file) else load_workbook(excel_file)
    ws = wb.active

    # Append the report data to the Excel file
    date = datetime.now().strftime("%Y-%m-%d")
    time_now = datetime.now().strftime("%H:%M:%S")
    report_data = [date, time_now, counter, total_time] + rep_times
    ws.append(report_data)

    # Save the changes to the Excel file
    wb.save(excel_file)

# Function to display report in GUI and save to excel file
def display_report_gui(counter, total_time, rep_times):
    root = tk.Tk()
    root.title("Deadlift Report")

    report_label = tk.Label(root, text="Report:", font=("Helvetica", 16, "bold"))
    report_label.pack(pady=10)

    counter_label = tk.Label(root, text=f"Counter: {counter}", font=("Helvetica", 14))
    counter_label.pack()

    total_time_label = tk.Label(root, text=f"Total Time: {round(total_time, 2)} sec", font=("Helvetica", 14))
    total_time_label.pack()

    rep_times_label = tk.Label(root, text="Individual Rep Times:", font=("Helvetica", 14))
    rep_times_label.pack()

    for idx, rep_time in enumerate(rep_times):
        rep_time_label = tk.Label(root, text=f"Rep {idx+1}: {round(rep_time, 2)} sec", font=("Helvetica", 14))
        rep_time_label.pack()

    # Button to save report data to Excel
    save_button = tk.Button(root, text="Save Report", command=lambda: save_report_to_excel(counter, total_time, rep_times))
    save_button.pack(pady=10)

    root.mainloop()

# Prompt user to input number of deadlifts
root = tk.Tk()
root.withdraw()
num_deadlifts = simpledialog.askinteger("Input", "Enter the number of deadlifts:")
num_deadlifts += 1  # Doing increment as index starts from 0

# Create or load Excel file
excel_file = "reports_summary.xlsx"
wb = Workbook() if not os.path.isfile(excel_file) else load_workbook(excel_file)
ws = wb.active
if ws.max_row == 1:  # If it's the first report, write headers
    ws.append(["Date", "Time", "Counter", "Total Time"] + [f"Rep {i}" for i in range(1, num_deadlifts)])
wb.save(excel_file)

# Other variables
counter = 0
stage = None
start_time = None
last_transition_time = None
rep_times = []
timer_started = False
timer_duration = 0

# Opening and setting up webcam
cap = cv2.VideoCapture(0)
cap.set(3, screen_width)
cap.set(4, screen_height)

# Setting up mediapipe
with mp_pose.Pose(min_detection_confidence=0.7, min_tracking_confidence=0.7) as pose:
    while cap.isOpened():
        ret, frame = cap.read()

        # Flip the frame horizontally
        frame = cv2.flip(frame, 1)

        # Recolour image as mediapipe accepts RGB only
        image_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        # Make detection
        results = pose.process(image_rgb)

        try:
            landmarks = results.pose_landmarks.landmark
            # 11 :  shoulder
            # 23 : hip
            # 25 : knee

            shoulder = [landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value].x,
                        landmarks[mp_pose.PoseLandmark.LEFT_SHOULDER.value].y]

            hip = [landmarks[mp_pose.PoseLandmark.LEFT_HIP.value].x,
                   landmarks[mp_pose.PoseLandmark.LEFT_HIP.value].y]

            knee = [landmarks[mp_pose.PoseLandmark.LEFT_KNEE.value].x,
                    landmarks[mp_pose.PoseLandmark.LEFT_KNEE.value].y]

            angle = calculateAngle(shoulder, hip, knee)

            if angle > 170:
                stage = "DOWN"
                if counter == 1 and not timer_started:  # Start timer on first rep
                    start_time = time.time()
                    timer_started = True
            elif angle < 105 and stage == "DOWN":
                stage = "UP"
                counter += 1
                if last_transition_time is not None:  # Calculate and store rep time
                    rep_times.append(time.time() - last_transition_time)
                last_transition_time = time.time()
                if counter == 1:  # Reset timer on first rep
                    start_time = time.time()
                    timer_duration = 0

                if counter == num_deadlifts:  # Stop the loop when desired number of deadlifts is reached
                    break

        except Exception as e:
            pass

        # Update timer if it has started
        if timer_started:
            timer_duration = time.time() - start_time

        # Displaying the frame with information
        if results.pose_landmarks:
            mp_drawing.draw_landmarks(frame, results.pose_landmarks, mp_pose.POSE_CONNECTIONS,
                                      mp_drawing.DrawingSpec(color=(255, 0, 255), thickness=2, circle_radius=2),
                                      mp_drawing.DrawingSpec(color=(255, 0, 0), thickness=2, circle_radius=2))

        # Displaying the counter
        cv2.putText(frame, f"Counter: {counter}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2,
                    cv2.LINE_AA)

        # Displaying the stage (UP or DOWN)
        cv2.putText(frame, f"Stage: {stage}", (10, 70), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2, cv2.LINE_AA)

        # Displaying the timer
        timer_text = f"Timer: {round(timer_duration, 2)} sec"
        cv2.putText(frame, timer_text, (10, 110), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 255), 2, cv2.LINE_AA)

        # Displaying the rep times
        if rep_times:
            rep_time_text = "Rep Times: " + ", ".join([f"{round(t, 2)} sec" for t in rep_times])
            cv2.putText(frame, rep_time_text, (10, 150), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 255), 2, cv2.LINE_AA)

        # Calculate and display angle between shoulder and hip
        calculateAndDisplayAngle(frame, shoulder, hip)

        # Display the frame
        cv2.imshow("Image FEED", frame)

        if cv2.waitKey(10) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

# Displaying the report in GUI
total_time = sum(rep_times) + timer_duration
display_report_gui(counter, total_time, rep_times)
